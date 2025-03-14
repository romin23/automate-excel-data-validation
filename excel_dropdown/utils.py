import os
from openpyxl import load_workbook
from openpyxl import Workbook as pywb  
from pprint import pprint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def copy_sheet(source_ws, target_wb):
    # Create a new sheet in the target workbook with the same name
    new_ws = target_wb.create_sheet(title=source_ws.title)
    # Copy all rows and columns
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()
    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        new_ws.merge_cells(str(merged_range))
    return new_ws


def get_dropdown_mappings(dropdown_ws):
    mappings = {}
    # Get header row (first row)
    headers = [cell.value for cell in dropdown_ws[1]]
    # Iterate over columns
    for col_idx, header in enumerate(headers, 1):
        if header is None:
            continue
        header_clean = header.strip().lower()
        # Collect dropdown values (skip header and empty)
        values = []
        for row in dropdown_ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell_val = row[0].value
            if cell_val not in (None, ""):
                values.append(str(cell_val))
        if values:
            mappings[header_clean] = values
    return mappings


from xlsxwriter.workbook import Workbook as xlwb
def apply_dropdowns_xlsxwriter(template_path, output_path, dropdown_mappings):
    # Load the template workbook data
    wb_openpyxl = load_workbook(template_path)
    ws_openpyxl = wb_openpyxl.active

    # Create new workbook with XlsxWriter
    wb_xlsxwriter = xlwb(output_path)
    ws_xlsxwriter = wb_xlsxwriter.add_worksheet(ws_openpyxl.title)

    # Copy data and merged cells from OpenPyXL template
    # -------------------------------------------------
    # Copy cell values
    for row in ws_openpyxl.iter_rows():
        for cell in row:
            ws_xlsxwriter.write(cell.row - 1, cell.column - 1, cell.value)

    # Copy merged cells
    for merged_range in ws_openpyxl.merged_cells.ranges:
        ws_xlsxwriter.merge_range(
            merged_range.min_row - 1,
            merged_range.min_col - 1,
            merged_range.max_row - 1,
            merged_range.max_col - 1,
            ""
        )

    # Configure dropdown mappings
    row_to_dropdown_map = {
        # Complete these mappings using dropdown_mappings keys
        "Client's Prior Placement": "Prior Placement",
        "Current Residential Setting": "Residential Setting",
        "Faison Department": "Faison Department",
        "Faison Program": "Faison Program",
        "Parent/Guardian Engagement Rating (5-pt)": "Rating Scales (5)",
        "Communication Level (10-pt)": "Rating Scales (10)",
        "Level of Support and Independence (10-pt)": "Rating Scales (10)",
        "Safety & Dangerousness Assessment (10-pt)": "Rating Scales (10)",
        "Readiness for Community-Based Learning (5-pt)": "Rating Scales (5)",
        "Aggression towards Caregivers and/or Staff":"interfering or problem behavior",
        "Aggression towards Peers":"interfering or problem behavior",
        "Disruptiveness or Property Destruction":"interfering or problem behavior",
        "Elopement or Wandering":"interfering or problem behavior",
        "Motor Stereotypy":"interfering or problem behavior",
        "Self-Injury or Self-Directed Harm":"interfering or problem behavior",
        "Vocal Stereotypy":"interfering or problem behavior",
        "Other Behavior Causing Dysfunction":"interfering or problem behavior",
        "Restrictiveness of Next Placement": "Restrictiveness of Next Placement",
        "Reason for Discharge": "Reason for Discharge",
        "Early Education Center - Next Placement Type": "EEC - Next Placement Type",
        "Schools - Diploma Status": "Schools - Diploma Status"
    }
    # Apply dropdown validations
    for row_idx, row in enumerate(ws_openpyxl.iter_rows(), start=0):
        if row[1].value is None:  # Column B (0-based index 1)
            continue
        
        row_label = str(row[1].value).strip().lower()
        for label_key, dd_key in row_to_dropdown_map.items():
            if row_label == label_key.lower():
                dropdown_options = dropdown_mappings.get(dd_key.lower(), [])
                if dropdown_options:
                    # Apply to columns C-I (XlsxWriter 0-based columns 2-8)
                    for col_idx in range(2, 9):
                        ws_xlsxwriter.data_validation(
                            row_idx, col_idx,
                            row_idx, col_idx,
                            {
                                'validate': 'list',
                                'source': dropdown_options,
                                'dropdown': True
                            }
                        )
                break

    # Add test dropdown to A1
    ws_xlsxwriter.data_validation(
        0, 0, 0, 0,
        {'validate': 'list', 'source': ['Test1', 'Test2'], 'dropdown': True}
    )

    wb_xlsxwriter.close()